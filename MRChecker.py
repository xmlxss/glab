import requests
from datetime import datetime, timedelta
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ---------------- CONFIG ----------------
GITLAB_API_URL = "https://gitlab.com/api/v4"
PROJECT_ID = "xxxx"
PRIVATE_TOKEN = "xxxx"
HOURS_THRESHOLD = 48
TIMEZONE = "Europe/Amsterdam" 
OUTPUT_FILE = "open_mrs.xlsx"
PER_PAGE = 100  
# ----------------------------------------

def is_weekend(date):
    return date.weekday() >= 5  # 5=zaterdag, 6=zondag

def working_hours_diff(start_time, end_time):
    """
    Calculate elapsed working hours between two datetimes excluding weekends
    """
    current = start_time
    hours = 0
    while current < end_time:
        if not is_weekend(current):
            next_hour = current + timedelta(hours=1)
            if next_hour > end_time:
                next_hour = end_time
            hours += (next_hour - current).total_seconds() / 3600
        current += timedelta(hours=1)
    return hours

def get_all_open_mrs():
    """
    Fetch all open merge requests.
    """
    mrs = []
    page = 1
    headers = {"PRIVATE-TOKEN": PRIVATE_TOKEN}

    while True:
        print(f"Fetching page {page} of open MRs...")
        params = {"state": "opened", "scope": "all", "per_page": PER_PAGE, "page": page}
        url = f"{GITLAB_API_URL}/projects/{PROJECT_ID}/merge_requests"
        response = requests.get(url, headers=headers, params=params)
        response.raise_for_status()
        page_mrs = response.json()
        if not page_mrs:
            break
        mrs.extend(page_mrs)
        page += 1

    print(f"Total open MRs fetched: {len(mrs)}")
    return mrs

def create_excel_report(mrs):
    tz = pytz.timezone(TIMEZONE)
    now = datetime.now(tz)

    wb = Workbook()
    ws = wb.active
    ws.title = "Open MRs"

    headers = ["MR Title", "Assignee", "Reviewers", "Created At", "Hours Waiting"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')

    for mr in mrs:
        created_at = datetime.strptime(mr['created_at'], "%Y-%m-%dT%H:%M:%S.%fZ")
        created_at = pytz.utc.localize(created_at).astimezone(tz)
        assignee = mr['assignee']['username'] if mr['assignee'] else "None"
        reviewers_list = [r['username'] for r in mr.get('reviewers', [])]
        reviewers = ", ".join(reviewers_list) if reviewers_list else "NO REVIEWERS ASSIGNED"
        hours_waiting = working_hours_diff(created_at, now)

        row = [mr['title'], assignee, reviewers, created_at.strftime("%Y-%m-%d %H:%M"), round(hours_waiting, 1)]
        ws.append(row)

        # Styling: No reviewers -> red, overdue -> yellow
        row_num = ws.max_row
        if not reviewers_list:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif hours_waiting > HOURS_THRESHOLD:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 5

    wb.save(OUTPUT_FILE)
    print(f"Excel report saved as '{OUTPUT_FILE}'")

if __name__ == "__main__":
    open_mrs = get_all_open_mrs()
    create_excel_report(open_mrs)